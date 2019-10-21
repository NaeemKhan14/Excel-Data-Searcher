from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, colors
import os


class Processor(object):
    # GLobal variables
    workbook = sheet = input_list = regex_list = None
    result_file = fname = ""
    found_results_list = []

    def __init__(_, name, input, regex, sheet_number=None):
        # Load the input file
        Processor.fname = name
        Processor.workbook = load_workbook(name)
        # Save the name of the file to later use for the new file
        Processor.result_file = os.path.basename(name).split(".")
        # Select the sheet number to work with. If nothing is given, default
        # first sheet is selected
        Processor.sheet = Processor.workbook[sheet_number] if sheet_number else Processor.workbook.active
        # Split the input (words to search for) into an array, and trim white spaces
        Processor.input_list = [x.strip() for x in input.lower().split(',')]
        # Split the regex input into an array, and trim white spaces
        Processor.regex_list = [x.strip() for x in regex.lower().split(',')]

    def process_rows(self, search_col, result_col):
        results_file_name = Processor.result_file[0] + "_RESULTS." + Processor.result_file[1]
        # If file doesn't exist, create it
        if not os.path.exists(results_file_name):
            self.create_new_file(results_file_name)
        # Load the new file to write results into
        new_file = load_workbook(results_file_name)
        sheet = new_file.active
        sheet.append(["-----------------------------------------------------------", ""])

        # Go through each row, and first check for the input (full matching word),
        # if it is found, put it in the new file. Otherwise check if the regex flag
        # is set. If it is, then check if this row (in the specified column) contains
        # that word, and add it in the list.
        for row in Processor.sheet:
            if Processor.input_list[0] != '' and row[search_col-1].value.lower() in Processor.input_list:
                self.write_to_file(row[result_col-1].value, sheet)
                row[search_col-1].fill = PatternFill(start_color=colors.GREEN, fill_type="solid")
            else:
                if Processor.regex_list[0] != '':
                    for word in Processor.regex_list:
                        if row[search_col-1].value.lower().find(word) != -1:  # if word is found
                            self.write_to_file(row[result_col-1].value, sheet)
                            row[search_col - 1].fill = PatternFill(start_color=colors.GREEN, fill_type="solid")

        # Save the file
        Processor.workbook.save(Processor.fname)
        new_file.save(filename=results_file_name)

    # Creates a new Excel file if it doesn't exist
    def create_new_file(self, filename):
        new_file = Workbook()
        new_sheet = new_file.active
        new_sheet.title = "Results sheet"
        new_file.save(filename=filename)

    def write_to_file(self, result_word, sheet):
        if result_word not in Processor.found_results_list:
            sheet.append([result_word, ""])
            Processor.found_results_list.append(result_word)
